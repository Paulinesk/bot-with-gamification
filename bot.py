import os
import sqlite3
import threading
from datetime import date, datetime
from random import choice
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import telebot
from telebot import types
from telebot.types import InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, ReplyKeyboardMarkup
from achievements import ACHIEVEMENTS

log_file_path = r'C:\Users\User\Desktop\hse_bot\bot.log'
logging.basicConfig(
    level=logging.INFO,  # Уровень логов: INFO — общая полезная информация
    format='%(asctime)s - %(levelname)s - %(message)s',  # Формат сообщения
    filename=log_file_path,  # Лог-файл для записи
    filemode='a'  # Режим
)
bot = telebot.TeleBot('')

ADMIN_PASSWORD = "HHtcALPg"
admin_sessions = set()
user_states = {}
user_current_task = {}  # user_id -> task_id
db_lock = threading.Lock()

def is_admin(user_id):
    return user_id in admin_sessions

def init_user_tasks(user_id):
    logging.info(f"init_user_tasks started for user_id={user_id}")
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()
    
        cur.execute('SELECT id FROM tasks')
        all_tasks = [row[0] for row in cur.fetchall()]
        logging.info(f"Total tasks found: {len(all_tasks)}")
    
        cur.execute('SELECT task_id FROM user_tasks WHERE user_id = ?', (user_id,))
        existing = set(row[0] for row in cur.fetchall())
        logging.info(f"User {user_id} already has {len(existing)} tasks")

        for task_id in all_tasks:
            if task_id not in existing:
                cur.execute('INSERT INTO user_tasks (user_id, task_id) VALUES (?, ?)', (user_id, task_id))
                logging.info(f"Added task_id={task_id} to user_id={user_id}")

        conn.commit()
        cur.close()
        conn.close()
    logging.info(f"init_user_tasks finished for user_id={user_id}")
def generate_excel_stat():
    logging.info("generate_excel_stat started")
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()

        cur.execute('''
        SELECT telegram_id, first_name, last_name, points, level FROM users
        ORDER BY points DESC
    ''')
        users = cur.fetchall()
        logging.info(f"Fetched {len(users)} users from database")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Статистика студентов"

        ws.append(["ID", "Имя", "Фамилия", "Баллы", "Уровень", "Выполнено заданий"])

        for user in users:
            telegram_id, first_name, last_name, points, level = user

        # Подсчёт правильно выполненных заданий
            cur.execute('SELECT COUNT(*) FROM user_tasks WHERE user_id = ? AND is_correct = 1', (telegram_id,))
            tasks_done = cur.fetchone()[0]

            ws.append([telegram_id, first_name, last_name, points, level, tasks_done])
            logging.info(f"Added stats for user_id={telegram_id}: tasks_done={tasks_done}")

        cur.close()
        conn.close()
        filename = "students_stats.xlsx"
        wb.save(filename)
        logging.info(f"Excel file saved as {filename}")
        return filename
def send_next_task(telegram_id, chat_id):
    logging.info(f"send_next_task called for user_id={telegram_id}")
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()
        cur.execute('''
            SELECT t.id, t.question, t.correct_answer
            FROM tasks t
            LEFT JOIN user_tasks ut ON t.id = ut.task_id AND ut.user_id = ?
            WHERE ut.completed IS NULL OR ut.completed = 0
            LIMIT 1
        ''', (telegram_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()

    if row:
        task_id, question, correct_answer = row
        user_current_task[telegram_id] = {
            "task_id": task_id,
            "correct_answer": correct_answer
        }
        logging.info(f"Sending task_id={task_id} to user_id={telegram_id}")
        bot.send_message(chat_id, f"📌 Новое задание:\n{question}")
    else:
        logging.info(f"No more tasks for user_id={telegram_id}")
        bot.send_message(chat_id, "🎉 Поздравляю! Все задания выполнены.")

@bot.message_handler(commands=['start'])
def start(message):
    logging.info(f"/start triggered by user_id={message.from_user.id}")
    try:
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()
            cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT,
            last_name TEXT,
            telegram_id INTEGER UNIQUE,
            points INTEGER DEFAULT 0,
            level INTEGER DEFAULT 1,
            role TEXT DEFAULT 'student'
        )
    ''')
            logging.info("Table 'users' ensured")
            cur.execute('''
    CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question TEXT,
        correct_answer TEXT,
        points INTEGER
    )''')
            logging.info("Table 'tasks' ensured")
            cur.execute('''
    CREATE TABLE IF NOT EXISTS user_tasks (
        user_id INTEGER,
        task_id INTEGER,
        completed INTEGER DEFAULT 0,
        is_correct INTEGER,
        PRIMARY KEY (user_id, task_id)
    )''')
            logging.info("Table 'user_tasks' ensured")
            cur.execute('''
    CREATE TABLE IF NOT EXISTS attendance_window (
        is_open INTEGER DEFAULT 0
    )''')
            logging.info("Table 'attendance_window' ensured")
            cur.execute('''
    CREATE TABLE IF NOT EXISTS attendance (
        user_id INTEGER,
        date TEXT,
        time TEXT,
        PRIMARY KEY (user_id, date)
    )''')
            logging.info("Table 'attendance' ensured")
            cur.execute('''
    CREATE TABLE IF NOT EXISTS achievements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE,
        name TEXT,
        description TEXT
    )''')
            logging.info("Table 'achievements' ensured")
            cur.execute('''
    CREATE TABLE IF NOT EXISTS user_achievements (
        user_id INTEGER,
        achievement_code TEXT,
        PRIMARY KEY (user_id, achievement_code)
    )''')
            logging.info("Table 'achievements' ensured")
            cur.execute('SELECT COUNT(*) FROM attendance_window')
            if cur.fetchone()[0] == 0:
               cur.execute('INSERT INTO attendance_window (is_open) VALUES (0)')
       
            conn.commit()
            cur.close()
            conn.close()
    except Exception as e:
        logging.error(f"Ошибка в /start: {e}", exc_info=True)
        bot.send_message(message.chat.id, "⚠️ Произошла ошибка при инициализации. Попробуйте позже.")
        return
    welcome_text = (
       "LevelUp Бот — твой персональный проводник в мире знаний и приключений!\n"
       "Прокачивай навыки, проходи задания, собирай достижения и соревнуйся с одногруппниками.Каждое задание — это шаг к новым уровням, баллам и званиям!\n"
       "👇 Для начала выбери свою роль:", 
    )
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("👨‍🎓 Студент", callback_data="role_student"))
    markup.add(InlineKeyboardButton("👩‍🏫 Преподаватель", callback_data="role_teacher"))

    bot.send_message(message.chat.id, welcome_text, reply_markup=markup)
    
def ask_last_name(message):
    telegram_id = message.from_user.id
    first_name = message.text.strip()
    logging.info(f"User {telegram_id} entered first name: {first_name}")
    if first_name.startswith('/'):
        bot.send_message(message.chat.id, "❗ Имя не должно начинаться с '/'. Попробуй ещё раз. Напиши имя (например, Иван):", parse_mode="Markdown")
        bot.register_next_step_handler(message, ask_last_name)
        return

    user_states[telegram_id]["first_name"] = first_name
    bot.send_message(message.chat.id, "Отлично! Теперь напиши *фамилию* (например, Петров):", parse_mode="Markdown")
    bot.register_next_step_handler(message, register_user)

def register_user(message):
    last_name = message.text.strip()
    if last_name.startswith('/'):
        bot.send_message(message.chat.id, "❗ Фамилия не должна начинаться с '/'. Попробуй ещё раз. Напиши *фамилию* (например, Петров):", parse_mode="Markdown")
        bot.register_next_step_handler(message, register_user)
        return
    user_id = message.from_user.id
    first_name = user_states[message.chat.id]['first_name']
    telegram_id = message.chat.id
    
    logging.info(f"Starting registration for: {first_name} {last_name} (telegram_id={telegram_id})")
    try:
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()
            cur.execute('INSERT OR IGNORE INTO users (first_name, last_name, telegram_id, role) VALUES (?, ?, ?, ?)', (first_name, last_name, telegram_id, 'student'))
            conn.commit()
            cur.execute('SELECT id FROM users WHERE telegram_id = ?', (telegram_id,))
            user_id = cur.fetchone()[0]
            cur.execute('INSERT OR IGNORE INTO user_achievements (user_id, achievement_code) VALUES (?, ?)', (user_id, 'registered'))
            conn.commit()
            cur.close()
            conn.close()
            logging.info(f"User registered successfully: id={user_id}, {first_name} {last_name}, telegram_id={telegram_id}")
    except Exception as e:
        logging.error(f"Registration error for telegram_id={telegram_id}: {e}", exc_info=True)
        bot.send_message(message.chat.id, "⚠️ Произошла ошибка во время регистрации. Попробуйте еще раз.")
        return
    main_menu = ReplyKeyboardMarkup(resize_keyboard=True)
    main_menu.row(KeyboardButton("📚 Задания"), KeyboardButton("📅 Посещаемость"))
    main_menu.row(KeyboardButton("👤 Профиль"), KeyboardButton("🏆 Лидерборд"))

    bot.send_message(message.chat.id, f"✅ Готово, {first_name} {last_name}! Теперь ты в игре 💪\n\n"
                                      "Введи /help, чтобы узнать, с чего начать 🚀",
                     reply_markup=main_menu)
def check_achievements(telegram_id):
    logging.info(f"Checking achievements for user {telegram_id}")
    try:
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()

            cur.execute('SELECT id, points, level FROM users WHERE telegram_id = ?', (telegram_id,))
            user_row = cur.fetchone()
            if not user_row:
                conn.close()
                return []

            user_id, points, level = user_row
            logging.info(f"User {telegram_id}: id={user_id}, points={points}, level={level}")

    # Получаем количество выполненных заданий
            cur.execute('SELECT COUNT(*) FROM user_tasks WHERE user_id = ? AND completed = 1', (user_id,))
            tasks_done = cur.fetchone()[0]

            cur.execute('SELECT achievement_code FROM user_achievements WHERE user_id = ?', (user_id,))
            existing = set(row[0] for row in cur.fetchall())
       
            new_achievements = []

            for code, name, description in ACHIEVEMENTS:
                if code in existing:
                    continue
                condition = (
                    (code == "first_10" and points >= 10) or
                    (code == "score_50" and points >= 50) or
                    (code == "score_150" and points >= 150) or
                    (code == "score_300" and points >= 300) or
                    (code == "lvl_2" and level >= 2) or
                    (code == "lvl_4" and level >= 4) or
                    (code == "lvl_6" and level >= 6) or
                    (code == "task_1" and tasks_done >= 1) or
                    (code == "task_10" and tasks_done >= 10) or
                    (code == "task_25" and tasks_done >= 25)
                )
                if condition:
                    cur.execute('INSERT OR IGNORE INTO user_achievements (user_id, achievement_code) VALUES (?, ?)', (user_id, code))
                    new_achievements.append(f"{name}: {description}")
            conn.commit()
            cur.close()
            conn.close()
            return new_achievements
    except Exception as e:
        logging.error(f"Error checking achievements for user {telegram_id}: {e}", exc_info=True)
        return []
@bot.callback_query_handler(func=lambda call: call.data.startswith('role_'))
def choose_role(call):
    bot.answer_callback_query(call.id)
    role = call.data.split('_')[1]
    user_id = call.from_user.id
    logging.info(f"User {user_id} selected role: {role}")
    user_states[call.from_user.id] = {'role': role}

    if role == 'student':
        bot.send_message(call.message.chat.id, "Как тебя зовут? Напиши *имя* (например, Иван):", parse_mode="Markdown")
        bot.register_next_step_handler(call.message, ask_last_name)  # продолжить регистрацию студента
    elif role == 'teacher':
        user_states[call.from_user.id]['awaiting_password'] = True
        bot.send_message(call.message.chat.id, "🔒 Введите пароль для входа в админ-панель:")
@bot.message_handler(commands=['admin'])
def admin_command(message):
    user_id = message.from_user.id
    user_states[message.from_user.id] = {'awaiting_password': True}
    logging.info(f"User {user_id} entered /admin command. Awaiting password input.")
    bot.send_message(message.chat.id, "🔒 Введите пароль для входа в админ-панель:")

@bot.message_handler(commands=['attend'])
def mark_attendance(message):
    user_id = message.from_user.id
    logging.info(f"[ATTEND] User {user_id} requested attendance")
    try:
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()

            cur.execute('SELECT is_open FROM attendance_window')
            is_open = cur.fetchone()[0]

            if not is_open:
                bot.send_message(message.chat.id, "❗ Сейчас окно посещаемости закрыто.")
                cur.close()
                conn.close()
                return

            today = date.today().isoformat()
            now_time = datetime.now().strftime("%H:%M:%S")

            cur.execute('SELECT * FROM attendance WHERE user_id = ? AND date = ?', (user_id, today))
            already_marked = cur.fetchone()

            if already_marked:
                bot.send_message(message.chat.id, "✅ Вы уже отметились сегодня.")
            else:
                cur.execute('INSERT INTO attendance (user_id, date, time) VALUES (?, ?, ?)', (user_id, today, now_time))
                points_for_attendance = 5
                cur.execute('UPDATE users SET points = points + ? WHERE telegram_id = ?', (points_for_attendance, user_id))

            # Обновляем уровень
                cur.execute('SELECT points FROM users WHERE telegram_id = ?', (user_id,))
                total_points = cur.fetchone()[0]
                new_level = total_points // 50 + 1
                cur.execute('UPDATE users SET level = ? WHERE telegram_id = ?', (new_level, user_id))

                conn.commit()
                logging.info(f"[ATTEND] User {user_id}: attendance marked at {now_time}, points={total_points}, level={new_level}")
                bot.send_message(message.chat.id, f"🟢 Посещение зафиксировано в {now_time}. Спасибо!")
    except Exception as e:
        logging.error(f"[ATTEND] Error for user {user_id}: {str(e)}")
        bot.send_message(message.chat.id, "⚠ Произошла ошибка при отметке посещения.")
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass
@bot.message_handler(func=lambda message: message.text == "📅 Посещаемость")
def handle_attendance_button(message):
    mark_attendance(message)

@bot.message_handler(commands=['leaderboard'])
def show_leaderboard(message):
    user_id = message.from_user.id
    logging.info(f"[LEADERBOARD] User {user_id} requested leaderboard")
    try:
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()

            cur.execute('''
        SELECT first_name, last_name, points
        FROM users
        ORDER BY points DESC
        LIMIT 5
        ''')
            top_users = cur.fetchall()
    except Exception as e:
        logging.error(f"[LEADERBOARD] Error for user {user_id}: {str(e)}")
        bot.send_message(message.chat.id, "⚠ Не удалось получить рейтинг.")
        return
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

    if not top_users:
        logging.info("[LEADERBOARD] Leaderboard is empty")
        bot.send_message(message.chat.id, "📉 Рейтинг пока пуст.")
        return

    text = "🏆 Топ-5 студентов:\n\n"
    for i, (first_name, last_name, points) in enumerate(top_users, start=1):
        name = f"{first_name or ''} {last_name or ''}".strip()
        text += f"{i}. {name} — {points} баллов\n"
        
    logging.info("[LEADERBOARD] Leaderboard sent to user")
    bot.send_message(message.chat.id, text)
@bot.message_handler(func=lambda message: message.text == "🏆 Лидерборд")
def handle_leaderboard_button(message):
    show_leaderboard(message)

@bot.message_handler(commands=['profile'])
def show_profile(message):
    telegram_id = message.from_user.id
    logging.info(f"User {telegram_id} requested profile")
    try:
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()

            cur.execute('''
        SELECT id, first_name, last_name, points, level
        FROM users
        WHERE telegram_id = ?
        ''', (telegram_id,))
            user_data = cur.fetchone()

            if not user_data:
                conn.close()
                bot.send_message(message.chat.id, "❌ Профиль не найден.")
                return

            user_id, first_name, last_name, points, level = user_data
            name = f"{first_name or ''} {last_name or ''}".strip()

            cur.execute('''
        SELECT achievement_code 
        FROM user_achievements 
        WHERE user_id = ?
        ''', (user_id,))
            ach_codes = [row[0] for row in cur.fetchall()]

            cur.close()
            conn.close()

        ach_names = [name for code, name, _ in ACHIEVEMENTS if code in ach_codes]
        achievements_text = "\n".join(ach_names) if ach_names else "— Нет достижений пока что —"

        profile_text = (
        f"👤 {name}\n"
        f"📈 Уровень: {level}   ⭐️ Баллы: {points}\n\n"
        f"🏅 Достижения:\n{achievements_text}"
        )
        bot.send_message(message.chat.id, profile_text)
    except Exception as e:
        logging.error(f"Error showing profile for user {telegram_id}: {e}", exc_info=True)
        bot.send_message(message.chat.id, "❗ Произошла ошибка при загрузке профиля. Попробуйте позже.")

@bot.message_handler(func=lambda message: message.text == "👤 Профиль")
def handle_profile_button(message):
    show_profile(message)
@bot.message_handler(commands=['help'])
def handle_help(message):
    help_text = (
        "Используй кнопки внизу экрана, чтобы:\n"
        "— проходить задания\n"
        "— смотреть свой профиль и достижения\n"
        "— проверять рейтинг\n"
        "— отмечать посещаемость(если окно посещаемости открыто преподавателем)\n\n"
        "❓ Если что-то не работает, напиши сюда: @paulinesssk"
    )
    bot.send_message(message.chat.id, help_text)

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('awaiting_password'))
def check_teacher_password(message):
    user_id = message.from_user.id
    if message.text.strip() == ADMIN_PASSWORD:
        admin_sessions.add(message.from_user.id)

        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()
            cur.execute('INSERT OR IGNORE INTO users (telegram_id, role) VALUES (?, ?)', (message.from_user.id, 'teacher'))
            cur.execute('UPDATE users SET role = ? WHERE telegram_id = ?', ('teacher', message.from_user.id))
            conn.commit()
            cur.close()
            conn.close()

        user_states.pop(message.from_user.id, None)
        logging.info(f"User {user_id} successfully logged in as teacher")

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add('📋 Список студентов', '📊 Статистика студентов')
        markup.add('🧩 Управление заданиями')
        markup.add('🟢 Открыть окно посещаемости')

        bot.send_message(message.chat.id, "✅ Доступ в админ-панель разрешен.", reply_markup=markup)
        
    else:
        logging.warning(f"User {user_id} entered incorrect admin password")
        bot.send_message(message.chat.id, "❌ Неверный пароль. Попробуйте снова.")
        
@bot.message_handler(func=lambda message: message.text == '📋 Список студентов')
def list_students(message):
    if not is_admin(message.from_user.id):
        return
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()

        cur.execute("SELECT first_name, last_name FROM users WHERE role = 'student' ORDER BY last_name ASC")
        students = cur.fetchall()
        
        if not students:
            bot.send_message(message.chat.id, "❗️Список студентов пуст.")
        else:
            response = "📋 *Список студентов:*\n\n"
            for i, (first_name, last_name) in enumerate(students, start=1):
                response += f"{i}. {last_name} {first_name}\n"

            bot.send_message(message.chat.id, response, parse_mode="Markdown")

        cur.close()
        conn.close()

@bot.message_handler(func=lambda message: message.text == '🧩 Управление заданиями')
def task_management_menu(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "❌ У вас нет доступа к этой функции.")
        return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('➕ Добавить задание', '📋 Список заданий')
    markup.add('🔙 Назад в главное меню')
    bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == '➕ Добавить задание')
def add_task_start(message):
    if not is_admin(message.from_user.id):
        return
    logging.info(f"Admin {message.from_user.id} started adding a new task.")
    user_states[message.from_user.id] = {'state': 'adding_question'}
    bot.send_message(message.chat.id, "Введите текст задания:")
    
@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'adding_question')
def add_task_question(message):
    logging.info(f"Admin {message.from_user.id} entered task question: {message.text.strip()}")
    user_states[message.from_user.id] = {
        'state': 'adding_answer',
        'question': message.text.strip()
    }
    bot.send_message(message.chat.id, "Введите правильный ответ:")
    bot.register_next_step_handler(message, add_task_answer)

def add_task_answer(message):
    user_data = user_states.get(message.from_user.id, {})
    if user_data.get('state') != 'adding_answer':
        logging.warning(f"User {message.from_user.id} sent answer in invalid state.")
        bot.send_message(message.chat.id, "❗️Произошла ошибка. Повторите добавление задания.")
        return

    user_data['answer'] = message.text.strip()
    user_data['state'] = 'adding_points'
    user_states[message.from_user.id] = user_data
    logging.info(f"User {message.from_user.id} entered answer: {user_data['answer']}")

    bot.send_message(message.chat.id, "Сколько баллов даётся за выполнение задания?")
    bot.register_next_step_handler(message, add_task_points)

def add_task_points(message):
    user_data = user_states.get(message.from_user.id, {})
    if user_data.get('state') != 'adding_points':
        logging.warning(f"User {message.from_user.id} sent points in invalid state.")
        bot.send_message(message.chat.id, "❗️Произошла ошибка. Повторите добавление задания.")
        return
    try:
        points = int(message.text.strip())
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()
            cur.execute('INSERT INTO tasks (question, correct_answer, points) VALUES (?, ?, ?)',
                        (user_data['question'], user_data['answer'], points))
            conn.commit()
            cur.close()
            conn.close()
        logging.info(f"User {message.from_user.id} added task: '{user_data['question']}' with {points} points.")
        bot.send_message(message.chat.id, "✅ Задание добавлено успешно!")
        user_states.pop(message.from_user.id, None)
    except ValueError:
        logging.warning(f"User {message.from_user.id} entered invalid points value: {message.text.strip()}")
        bot.send_message(message.chat.id, "❗ Введите целое число для баллов.")
        user_data['state'] = 'adding_points'
        user_states[message.from_user.id] = user_data
        bot.register_next_step_handler(message, add_task_points)

@bot.message_handler(func=lambda message: message.text == '📋 Список заданий')
def list_tasks(message):
    if not is_admin(message.from_user.id):
        logging.warning(f"User {message.from_user.id} tried to access task list without permission.")
        return
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()
        cur.execute('SELECT id, question, points FROM tasks')
        tasks = cur.fetchall()
        cur.close()
        conn.close()
    logging.info(f"User {message.from_user.id} requested task list.")
    
    if not tasks:
        bot.send_message(message.chat.id, "Нет заданий.")
        return

    for task in tasks:
        task_id, question, points = task
        text = f"📌 *#{task_id}* — {question}\n💠 Баллы: {points}"
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✏️ Редактировать", callback_data=f"edit_{task_id}"),
            types.InlineKeyboardButton("🗑 Удалить", callback_data=f"delete_{task_id}")
        )
        bot.send_message(message.chat.id, text, parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("delete_"))
def delete_task(call):
    task_id = int(call.data.split('_')[1])
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()
        cur.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
        conn.commit()
        cur.close()
        conn.close()
    logging.info(f"User {call.from_user.id} deleted task with ID {task_id}")
    bot.edit_message_text("❌ Задание удалено.", chat_id=call.message.chat.id, message_id=call.message.message_id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("edit_"))
def edit_task_start(call):
    task_id = int(call.data.split('_')[1])
    user_states[call.from_user.id] = {'state': 'editing_question', 'task_id': task_id}
    logging.info(f"User {call.from_user.id} started editing task ID {task_id}")
    bot.send_message(call.message.chat.id, "Введите новый текст задания:")

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'editing_question')
def edit_task_question(message):
    user_data = user_states.get(message.from_user.id, {})
    user_data['question'] = message.text.strip()
    user_data['state'] = 'editing_answer'
    user_states[message.from_user.id] = user_data
    logging.info(f"User {message.from_user.id} entered new question text for task ID {user_data['task_id']}")
    bot.send_message(message.chat.id, "Введите новый правильный ответ:")

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'editing_answer')
def edit_task_answer(message):
    user_data = user_states.get(message.from_user.id, {})
    user_data['answer'] = message.text.strip()
    user_data['state'] = 'editing_points'
    user_states[message.from_user.id] = user_data
    logging.info(f"User {message.from_user.id} entered new answer for task ID {user_data['task_id']}")
    bot.send_message(message.chat.id, "Введите новое количество баллов:")

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'editing_points')
def edit_task_points(message):
    user_data = user_states.get(message.from_user.id, {})
    try:
        points = int(message.text.strip())
        with db_lock:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()
            cur.execute('UPDATE tasks SET question=?, correct_answer=?, points=? WHERE id=?',
                        (user_data['question'], user_data['answer'], points, user_data['task_id']))
            conn.commit()
            cur.close()
            conn.close()
        logging.info(f"User {message.from_user.id} updated task ID {user_data['task_id']}")
        bot.send_message(message.chat.id, "✅ Задание успешно обновлено!")
        user_states.pop(message.from_user.id, None)
    except ValueError:
        logging.warning(f"User {message.from_user.id} entered invalid points value during editing: {message.text.strip()}")
        bot.send_message(message.chat.id, "❗ Введите целое число для баллов.")

@bot.message_handler(func=lambda message: message.text == "📊 Статистика студентов")
def student_statistics(message):
    if not is_admin(message.from_user.id):
        logging.warning(f"Unauthorized user {message.from_user.id} attempted to access statistics.")
        bot.reply_to(message, "Доступ запрещён.")
        return

    file_path = generate_excel_stat()
    logging.info(f"User {message.from_user.id} requested student statistics.")
    with open(file_path, 'rb') as file:
        bot.send_document(message.chat.id, file)

@bot.message_handler(func=lambda message: message.text == '🟢 Открыть окно посещаемости')
def open_attendance_window(message):
    if not is_admin(message.from_user.id):
        logging.warning(f"User {message.from_user.id} tried to open attendance window without permission.")
        return
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()
        cur.execute('UPDATE attendance_window SET is_open = 1')
        conn.commit()
        cur.close()
        conn.close()
    logging.info(f"Admin {message.from_user.id} opened attendance window.")
    # Создаём кнопку
    markup = InlineKeyboardMarkup()
    close_button = InlineKeyboardButton("🔴 Закрыть окно посещаемости", callback_data="close_attendance")
    markup.add(close_button)

    bot.send_message(message.chat.id, "✅ Окно посещаемости открыто.", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == 'close_attendance')
def close_attendance_window(call):
    bot.answer_callback_query(call.id)
    if not is_admin(call.from_user.id):
        logging.warning(f"User {call.from_user.id} tried to close attendance window without permission.")
        return
    today = date.today().isoformat()
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()

        cur.execute('UPDATE attendance_window SET is_open = 0')
        conn.commit()

        cur.execute('''
        SELECT u.first_name, u.last_name, u.telegram_id, a.time
        FROM attendance a
        JOIN users u ON a.user_id = u.telegram_id
        WHERE a.date = ?
        ''', (today,))
        students = cur.fetchall()

        cur.close()
        conn.close()
    logging.info(f"Admin {call.from_user.id} closed attendance window. {len(students)} students marked as present.")
    bot.send_message(call.message.chat.id, "⛔ Окно посещаемости закрыто.")

    if not students:
        bot.send_message(call.message.chat.id, "Сегодня никто не отметился.")
        return

    filename = f"attendance_{today}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    headers = ["№", "Фамилия Имя", "Telegram ID", "Время отметки"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, (first_name, last_name, telegram_id, time) in enumerate(students, 1):
        full_name = f"{last_name} {first_name}"
        ws.append([i, full_name, f"@{telegram_id}", time])

    wb.save(filename)

    with open(filename, 'rb') as file:
        bot.send_document(call.message.chat.id, file)

    os.remove(filename)
    logging.info(f"Attendance file {filename} sent and deleted.")
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()
        cur.execute('DELETE FROM attendance WHERE date = ?', (today,))
        conn.commit()
        cur.close()
        conn.close()
    logging.info(f"Attendance data for {today} deleted from database.")
    
@bot.message_handler(func=lambda message: message.text == '🔙 Назад в главное меню')
def back_to_main_menu(message):
    if not is_admin(message.from_user.id):
        logging.warning(f"User {message.from_user.id} tried to access admin menu.")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('📋 Список студентов', '📊 Статистика студентов')
    markup.add('🧩 Управление заданиями')
    markup.add('🟢 Открыть окно посещаемости')

    bot.send_message(message.chat.id, "Вы вернулись в главное меню.", reply_markup=markup)
    logging.info(f"Admin {message.from_user.id} returned to main menu.")

@bot.message_handler(commands=['tasks'])
def handle_tasks(message):
    user_id = message.from_user.id
    logging.info(f"User {user_id} requested /tasks.")
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()

        cur.execute('SELECT id FROM tasks')
        all_task_ids = [row[0] for row in cur.fetchall()]
        logging.debug(f"All task IDs: {all_task_ids}")

        cur.execute('SELECT task_id FROM user_tasks WHERE user_id = ?', (user_id,))
        existing_task_ids = set(row[0] for row in cur.fetchall())
        logging.debug(f"Existing tasks for user {user_id}: {existing_task_ids}")

        for task_id in all_task_ids:
            if task_id not in existing_task_ids:
                cur.execute(
                    'INSERT INTO user_tasks (user_id, task_id, completed) VALUES (?, ?, 0)',
                    (user_id, task_id)
                )
        conn.commit()

        cur.execute('''
        SELECT t.id, t.question, t.correct_answer
        FROM tasks t
        JOIN user_tasks ut ON t.id = ut.task_id
        WHERE ut.user_id = ? AND ut.completed = 0
        ''', (user_id,))
        rows = cur.fetchall()
        cur.close()
        conn.close()
    if not rows:
        logging.info(f"User {user_id} has completed all tasks.")
        bot.send_message(message.chat.id, "🎉 Все задания выполнены!")
        conn.close()
        return

    task = choice(rows)
    task_id, task_question, correct_answer = task

    user_current_task[user_id] = {
        "task_id": task_id,
        "correct_answer": correct_answer
    }
    logging.info(f"Task {task_id} sent to user {user_id}.")
    bot.send_message(message.chat.id, f"📌 Задание:\n{task_question}")
@bot.message_handler(func=lambda message: message.text == "📚 Задания")
def handle_tasks_button(message):
    handle_tasks(message)

@bot.callback_query_handler(func=lambda call: call.data == "next_task")
def handle_next_task(call):
    message = call.message
    telegram_id = call.from_user.id
    bot.answer_callback_query(call.id) #check
    logging.info(f"User {telegram_id} requested next task via inline button.")
    with db_lock:
        conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
        cur = conn.cursor()

        cur.execute('''
        SELECT t.id, t.question, t.correct_answer
        FROM tasks t
        LEFT JOIN user_tasks ut ON t.id = ut.task_id AND ut.user_id = ?
        WHERE ut.completed IS NULL OR ut.completed = 0
        LIMIT 1
        ''', (telegram_id,))
        task = cur.fetchone()

        cur.close()
        conn.close()

    if not task:
        logging.info(f"User {telegram_id} has no remaining tasks.")
        bot.send_message(message.chat.id, "🎉 Все задания выполнены!")
        return

    task_id, question, answer = task
    user_current_task[telegram_id] = {"task_id": task_id, "correct_answer": answer}
    logging.info(f"Sent next task {task_id} to user {telegram_id}.")
    bot.send_message(message.chat.id, f"🧩 Задание:\n\n{question}")

@bot.message_handler(func=lambda message: message.from_user.id in user_current_task)
def handle_task_answer(message):
    telegram_id = message.from_user.id
    user_answer = message.text.strip().lower()
    task_info = user_current_task[telegram_id]
    task_id = task_info["task_id"]
    correct_answer = task_info["correct_answer"].strip().lower()

    is_correct = user_answer == correct_answer
    logging.debug(f"[Ответ] Пользователь {telegram_id}: '{user_answer}' vs '{correct_answer}' => {'верно' if is_correct else 'неверно'}")
    points = 0
    total_points = None
    new_level = None
    new_ach = []

    with db_lock:
        try:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()

            # Получаем количество баллов за задание
            cur.execute('SELECT points FROM tasks WHERE id = ?', (task_id,))
            result = cur.fetchone()
            points = result[0] if result else 0

            # Помечаем задание как выполненное (даже если неверно)
            cur.execute('''
                INSERT INTO user_tasks (user_id, task_id, completed, is_correct)
                VALUES (?, ?, 1, ?)
                ON CONFLICT(user_id, task_id)
                DO UPDATE SET completed = 1, is_correct = excluded.is_correct
            ''', (telegram_id, task_id, int(is_correct)))


            if is_correct:
                cur.execute('''
                    UPDATE users
                    SET points = points + ?
                    WHERE telegram_id = ?
                ''', (points, telegram_id))

                cur.execute('SELECT points FROM users WHERE telegram_id = ?', (telegram_id,))
                total_points = cur.fetchone()[0]
                new_level = total_points // 50 + 1

                cur.execute('UPDATE users SET level = ? WHERE telegram_id = ?', (new_level, telegram_id))
                logging.debug(f"[Баллы] Задание {task_id}, начислено: {points}, всего: {total_points}, уровень: {new_level}")

            conn.commit()

        except Exception as e:
            logging.exception("Ошибка при проверке задания")

        finally:
            cur.close()
            conn.close()

    if is_correct:
        new_ach = check_achievements(telegram_id)

    with db_lock:
        try:
            conn = sqlite3.connect(r'C:\Users\User\Desktop\hse_bot\levelupbot.db', timeout=30)
            cur = conn.cursor()
            cur.execute('''
                SELECT COUNT(*) FROM user_tasks
                WHERE user_id = ? AND completed = 0
            ''', (telegram_id,))
            remaining = cur.fetchone()[0]
            logging.debug(f"[Прогресс] Пользователь {telegram_id} — осталось заданий: {remaining}")
        finally:
            cur.close()
            conn.close()

    if is_correct:
        msg = f"✅ Верно! +{points} баллов.\n\n📌 Осталось заданий: {remaining}"
        if new_ach:
            msg += "\n\n🏅 Новые достижения:\n" + "\n".join(new_ach)
    else:
        msg = f"❌ Неверно.\n📌 Осталось заданий: {remaining}"

    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➡️ Дальше", callback_data="next_task"))
    bot.send_message(message.chat.id, msg, reply_markup=keyboard)
    
    user_current_task.pop(telegram_id, None)

@bot.message_handler(func=lambda message: True)
def handle_unknown(message):
    logging.info(f"[Неизвестное сообщение] От {message.from_user.id}: {message.text}")
    bot.send_message(
        message.chat.id,
        "🤔 Я пока не понимаю такие сообщения.\nПожалуйста, используй кнопки ниже или команду /help."
    )    
if __name__ == '__main__':
    bot.polling(none_stop=True)
